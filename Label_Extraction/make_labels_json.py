#!/usr/bin/env python3
"""
FBSC Excel -> JSON for sheets 1.0..7.0 using fixed column ORDER.

Column order (1-indexed):
  1: ID
  2: Observation       (label)
  3: Cause/Effect
  4: choice_order      (a/b/c/d, one per row)
  5: Choice            (short label for the recommendation option)
  6: Text              (long recommendation paragraph)

USAGE:
  Place this file next to:
    FBSC_Roof_Assessment_Report_Obs_Type.xlsx
  Then run:
    python make_roof_templates_all_sheets.py

OUTPUT:
  forsmith_roof_labels.json  (aggregated across sheets 1.0..7.0)

Notes:
- Strictly positional columns; no column-name guessing or swapping.
- Preserves internal newlines in recommendation text.
- Handles merged ID blocks and collects ALL options within each block.
- Normalizes column 4 strictly to single letters (a..z); if blank, auto-assigns a,b,c,...
- Items are grouped by (sheet_name, id, label, cause_effect) so same IDs on different sheets do not collide.
"""

from __future__ import annotations
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

EXCEL_FILENAME = "FBSC_Roof_Assessment_Report_Obs_Type.xlsx"
OUT_JSON       = "forsmith_roof_labels.json"

# Sheets we want (by friendly targets). We'll try exact matches,
# then prefix matches like "1.0 ..." if needed.
TARGET_SHEETS = ["1.0", "2.0", "3.0", "4.0", "5.0", "6.0", "7.0"]

# 0-indexed column positions (strict by order)
COL_ID              = 0  # 1: ID
COL_OBSERVATION     = 1  # 2: Observation
COL_CAUSE_EFFECT    = 2  # 3: Cause/Effect
COL_CHOICE_LETTER   = 3  # 4: choice_order (a/b/c/d)
COL_REC_NAME        = 4  # 5: Choice (short label)
COL_REC_TEXT        = 5  # 6: Text (long paragraph)

# ---------------------------- Excel loaders ----------------------------

def load_sheet_pandas(path: Path, sheet_name: str):
    import pandas as pd
    df = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=str)
    df = df.fillna("")
    rows = [[str(v) if v is not None else "" for v in row] for row in df.values.tolist()]
    # strip ends but preserve internal newlines
    return [[c.strip() if isinstance(c, str) else "" for c in row] for row in rows]

def load_sheet_openpyxl(wb, sheet_name: str):
    ws = wb[sheet_name]
    rows = []
    for r in ws.iter_rows(values_only=True):
        row = []
        for v in r:
            if v is None:
                row.append("")
            else:
                s = str(v)
                row.append(s.strip())  # keep internal newlines, strip ends
        rows.append(row)
    return rows

def list_sheets_pandas(path: Path):
    import pandas as pd
    return pd.ExcelFile(path).sheet_names

def list_sheets_openpyxl(path: Path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    return wb, wb.sheetnames

# ---------------------------- Utilities ----------------------------

def pick_actual_sheet_names(available: list[str]) -> list[str]:
    """Map TARGET_SHEETS to actual sheet names.
       Tries exact match, then 'starts with' match (case-insensitive).
    """
    chosen = []
    low_map = {s.lower(): s for s in available}

    for target in TARGET_SHEETS:
        t_low = target.lower()
        # exact
        if t_low in low_map:
            chosen.append(low_map[t_low])
            continue
        # startswith (e.g., "1.0 - Roofing")
        found = None
        for s in available:
            if s.lower().startswith(t_low):
                found = s
                break
        if found:
            chosen.append(found)
        else:
            print(f"⚠️  Warning: target sheet '{target}' not found.")
    return chosen

def maybe_drop_header(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return rows
    header_words = {"id", "observation", "cause", "cause/effect", "choice", "choice_order", "recommendation", "text", "recommendation text"}
    first = [c.lower() for c in rows[0]]
    if any(any(h in cell for h in header_words) for cell in first):
        return rows[1:]
    return rows

def id_key_for_sort(s: str):
    parts = []
    for p in s.split("."):
        try:
            parts.append(int(p))
        except ValueError:
            parts.append(p)
    return tuple(parts)

def sheet_key_for_sort(s: str):
    """Sort sheets naturally by numeric prefix if present (e.g., '1.0', '2.0 - x')."""
    m = re.match(r"^\s*(\d+)(?:\.(\d+))?", s)
    if not m:
        return (9999, 9999, s.lower())
    major = int(m.group(1))
    minor = int(m.group(2)) if m.group(2) else 0
    return (major, minor, s.lower())

def norm_letter_strict(raw: str) -> str | None:
    """
    STRICT: return a single lowercase letter a..z only if the trimmed value is exactly one letter
    or one letter + simple trailing punctuation (e.g., 'a', 'a)', 'a.', '(a)' -> 'a').
    Otherwise return None (caller will auto-assign).
    """
    if not raw:
        return None
    t = raw.strip().lower()
    # remove simple leading/trailing punctuation/brackets
    t = re.sub(r"^[\(\[\{<\s]+", "", t)
    t = re.sub(r"[\)\]\}>\s\.:;\-]+$", "", t)
    if len(t) == 1 and t.isalpha():
        return t
    return None

# ---------------------------- Core ----------------------------

def main():
    here = Path.cwd()
    xlsx = here / EXCEL_FILENAME
    if not xlsx.exists():
        print(f"❌ Excel not found next to this script: {xlsx}")
        print("   Move this script next to the Excel OR edit EXCEL_FILENAME at the top.")
        sys.exit(1)

    # Discover sheets (pandas → openpyxl)
    used_backend = None
    all_sheets = []
    wb = None
    try:
        all_sheets = list_sheets_pandas(xlsx)
        used_backend = "pandas"
    except Exception:
        try:
            wb, all_sheets = list_sheets_openpyxl(xlsx)
            used_backend = "openpyxl"
        except Exception as e:
            print("❌ Could not list sheets (pandas or openpyxl). Try: pip install pandas openpyxl")
            print("   Error:", e)
            sys.exit(2)

    target_actual = pick_actual_sheet_names(all_sheets)
    if not target_actual:
        print("❌ None of the target sheets 1.0..7.0 were found.")
        print("   Available sheets:", all_sheets)
        sys.exit(3)

    grouped = {}  # key: (sheet_name, id, label, cause) -> list of rec rows

    for sheet in target_actual:
        # Load rows
        if used_backend == "pandas":
            rows = load_sheet_pandas(xlsx, sheet)
        else:
            rows = load_sheet_openpyxl(wb, sheet)
        rows = maybe_drop_header(rows)

        # --- Walk rows and build merged ID blocks; collect ALL options in each block ---
        i = 0
        n = len(rows)
        while i < n:
            row = rows[i] + [""] * max(0, 6 - len(rows[i]))
            uid   = row[COL_ID].strip()
            label = row[COL_OBSERVATION].strip()
            cause = row[COL_CAUSE_EFFECT].strip()

            # Skip empty lines
            if not any(row[:6]):
                i += 1
                continue

            # If not a new block (no ID/label), move on
            if not uid or not label:
                i += 1
                continue

            # Find the extent of this merged ID block:
            j = i + 1
            while j < n:
                nxt = rows[j] + [""] * max(0, 6 - len(rows[j]))
                nxt_uid = nxt[COL_ID].strip()
                nxt_label = nxt[COL_OBSERVATION].strip()
                # A new block starts when a new UID appears (merged cell ends)
                if nxt_uid:
                    break
                j += 1

            block = [rows[k] + [""] * max(0, 6 - len(rows[k])) for k in range(i, j)]

            # ---- STRICT letter handling (col 4), with sequential fallback ----
            auto_letters = [chr(ord('a') + k) for k in range(len(block))]
            block_letters = []
            for k, r in enumerate(block):
                L = norm_letter_strict(r[COL_CHOICE_LETTER])
                block_letters.append(L if L else auto_letters[k])

            # ---- Choice (col 5) and Text (col 6) used AS-IS (no swapping) ----
            block_choice_names = [(r[COL_REC_NAME] or "").strip() for r in block]
            block_texts        = [(r[COL_REC_TEXT] or "").strip() for r in block]

            # Accumulate for this group
            key = (sheet, uid, label, cause)
            for k in range(len(block)):
                grouped.setdefault(key, []).append({
                    "letter": block_letters[k],
                    "choice": block_choice_names[k],
                    "text": block_texts[k]
                })

            i = j  # next block

    # Build items
    items = []
    for (sheet, uid, label, cause), recs in grouped.items():
        choice_order = []
        rec_map = {}
        for x in recs:
            L = x["letter"].lower()
            if L not in choice_order:
                choice_order.append(L)
            rec_map[L] = {"choice": x["choice"], "text": x["text"]}

        items.append({
            "sheet": sheet,
            "id": uid,
            "label": label,
            "cause_effect": cause,
            "choice_order": choice_order,
            "recommendations": rec_map
        })

    # Sort first by sheet (natural), then by id (natural)
    items.sort(key=lambda it: (sheet_key_for_sort(it["sheet"]), id_key_for_sort(it["id"])))

    out = {
        "version": 1,
        "source_sheets": target_actual,
        "items": items
    }

    Path(OUT_JSON).write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")

    # Per-sheet breakdown
    per_sheet_counts = defaultdict(int)
    for it in items:
        per_sheet_counts[it["sheet"]] += 1

    sheets_sorted = sorted(per_sheet_counts.keys(), key=sheet_key_for_sort)
    total_items = len(items)

    print("✅ Done.")
    print(f"   Backend       : {used_backend}")
    print(f"   Sheets found  : {len(all_sheets)} → used: {target_actual}")
    print(f"   Items written : {total_items}")
    print(f"   Output        : {OUT_JSON}")
    print("   Breakdown by sheet:")
    for s in sheets_sorted:
        print(f"     {s}: {per_sheet_counts[s]} labels")

if __name__ == "__main__":
    main()
